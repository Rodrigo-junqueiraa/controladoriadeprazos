import sys
import os
import sys
import tkinter as tk
from tkinter import messagebox, filedialog, ttk
from openpyxl import load_workbook
from PIL import Image, ImageTk
from datetime import datetime, timedelta
import numpy as np
import pandas as pd
import json
from datetime import datetime

import sys
import zipfile
from datetime import datetime

def abrir_janela_preenchimento_multiplo():
    nova_janela = tk.Toplevel(janela)
    nova_janela.title("Múltiplos Preenchimentos")
    nova_janela.geometry("520x720")
    nova_janela.configure(bg="#0f172a")
    
    

    def criar_label(texto):
        return tk.Label(nova_janela, text=texto, bg="#0f172a", fg="white")

    criar_label("Data do Prazo (DD/MM):").pack()
    entry_data_prazo = tk.Entry(nova_janela)
    entry_data_prazo.pack()

    criar_label("Nome:").pack()
    entry_nome = tk.Entry(nova_janela)
    entry_nome.pack()

    criar_label("Processo:").pack()
    entry_processo = tk.Entry(nova_janela)
    entry_processo.pack()

    criar_label("Tipo de Prazo:").pack()
    entry_tipo = tk.Entry(nova_janela)
    entry_tipo.pack()

    criar_label("Responsável:").pack()
    entry_resp = tk.Entry(nova_janela)
    entry_resp.pack()

    criar_label("Data para notificar o Fatal (DD/MM):").pack()
    entry_data_notificar = tk.Entry(nova_janela)
    entry_data_notificar.pack()

    criar_label("Publicação:").pack()
    entry_pub = tk.Entry(nova_janela)
    entry_pub.pack()

    criar_label("Observações:").pack()
    entry_obs = tk.Text(nova_janela, height=3, wrap="word")
    entry_obs.pack(fill="x", padx=10, pady=5)

    colunas = ("Nome", "Processo", "Tipo", "Data", "Resp", "Pub", "Obs")
    tree = ttk.Treeview(nova_janela, columns=colunas, show="headings")
    for col in colunas:
        tree.heading(col, text=col)
    tree.pack(pady=5, fill="both", expand=True)

    dado_editando = None

    def adicionar():
        nome = entry_nome.get()
        processo = entry_processo.get()
        tipo = entry_tipo.get()
        data = entry_data_prazo.get()
        resp = entry_resp.get()
        pub = entry_pub.get()
        obs = entry_obs.get("1.0", tk.END).strip()
        data_notificar = entry_data_notificar.get()
        if all([nome, processo, tipo, data, data_notificar, resp, pub]):
            for r in registros:
                if r[0] == nome and r[1] == processo and r[2] == tipo:
                    messagebox.showwarning("Duplicidade", "Este registro já existe!")
                    return
            registros.append((nome, processo, tipo, data, data_notificar, resp, pub, obs))
            tree.insert("", "end", values=(nome, processo, tipo, data, resp, pub, obs[:30] + ("..." if len(obs) > 30 else "")))
            entry_nome.delete(0, tk.END)
            entry_processo.delete(0, tk.END)
            entry_tipo.delete(0, tk.END)
            entry_data_prazo.delete(0, tk.END)
            entry_resp.delete(0, tk.END)
            entry_pub.delete(0, tk.END)
            entry_obs.delete("1.0", tk.END)

    def editar_registro():
        nonlocal dado_editando
        item = tree.selection()
        if not item:
            messagebox.showwarning("Aviso", "Selecione um registro para editar.")
            return
        dado_editando = item[0]
        valores = tree.item(dado_editando, "values")
        entry_nome.delete(0, tk.END)
        entry_nome.insert(0, valores[0])
        entry_processo.delete(0, tk.END)
        entry_processo.insert(0, valores[1])
        entry_tipo.delete(0, tk.END)
        entry_tipo.insert(0, valores[2])
        entry_data_prazo.delete(0, tk.END)
        entry_data_prazo.insert(0, valores[3])
        entry_resp.delete(0, tk.END)
        entry_resp.insert(0, valores[4])
        entry_pub.delete(0, tk.END)
        entry_pub.insert(0, valores[5])
        entry_obs.delete("1.0", tk.END)
        entry_obs.insert("1.0", valores[6])

    def salvar_alteracoes():
        nonlocal dado_editando
        if not dado_editando:
            messagebox.showwarning("Aviso", "Nenhum registro selecionado para edição.")
            return
        nome = entry_nome.get()
        processo = entry_processo.get()
        tipo = entry_tipo.get()
        data = entry_data_prazo.get()
        resp = entry_resp.get()
        pub = entry_pub.get()
        obs = entry_obs.get("1.0", tk.END).strip()
        data_notificar = entry_data_notificar.get()
        indice = tree.index(dado_editando)
        registros[indice] = (nome, processo, tipo, data, data_notificar, resp, pub, obs)
        tree.item(dado_editando, values=(nome, processo, tipo, data, resp, pub, obs[:30] + ("..." if len(obs) > 30 else "")))
        dado_editando = None

    def excluir_registro():
        item = tree.selection()
        if not item:
            messagebox.showwarning("Aviso", "Selecione um registro para excluir.")
            return
        confirmar = messagebox.askyesno("Confirmação", "Deseja excluir este registro?")
        if confirmar:
            indice = tree.index(item)
            registros.pop(indice)
            tree.delete(item)

    def gerar_excel():
        if not registros:
            messagebox.showerror("Erro", "Nenhum dado para preencher.")
            return
        try:
            wb = load_workbook(ARQUIVO_MODELO)
            ws = wb.active
            bloco_index = 0
            for i, row in enumerate(ws.iter_rows(min_row=1)):
                if bloco_index >= len(registros):
                    break
                if row[0].value and str(row[0].value).strip().upper() == "NOME" and row[1].value in (None, ""):
                    nome, processo, tipo, data, data_notificar, resp, pub, obs = registros[bloco_index]
                    ws.cell(row=i+1, column=2).value = nome
                    ws.cell(row=i+2, column=2).value = processo
                    ws.cell(row=i+3, column=2).value = tipo
                    ws.cell(row=i+1, column=4).value = resp
                    ws.cell(row=i+2, column=4).value = pub
                    ws.cell(row=i+3, column=4).value = data
                    bloco_index += 1
            salvar_em = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel", "*.xlsx")])
            if salvar_em:
                wb.save(salvar_em)
                registrar_prazos_em_json()
                messagebox.showinfo("Sucesso", f"Arquivo salvo em: {salvar_em}")
        except Exception as e:
            messagebox.showerror("Erro ao preencher planilha", str(e))


    # Frame para agrupar os botões em linha
    frame_botoes = tk.Frame(nova_janela, bg="#0f172a")
    frame_botoes.pack(pady=10)

    btn_adicionar = tk.Button(frame_botoes, text="Adicionar", bg="#2563eb", fg="white", command=adicionar)
    btn_adicionar.original_bg = "#2563eb"
    configurar_estilo_botoes(btn_adicionar)
    btn_adicionar.bind("<Enter>", efeito_hover)
    btn_adicionar.bind("<Leave>", efeito_sair_hover)
    btn_adicionar.pack(side="left", padx=5)

    btn_editar = tk.Button(frame_botoes, text="Editar Registro", bg="#2563eb", fg="white", command=editar_registro)
    btn_editar.original_bg = "#2563eb"

    configurar_estilo_botoes(btn_editar)
    btn_editar.bind("<Enter>", efeito_hover)
    btn_editar.bind("<Leave>", efeito_sair_hover)
    btn_editar.pack(side="left", padx=5)

    btn_salvar = tk.Button(frame_botoes, text="Salvar Alterações", bg="#2563eb", fg="white", command=salvar_alteracoes)
    btn_salvar.original_bg = "#2563eb"
    configurar_estilo_botoes(btn_salvar)
    btn_salvar.bind("<Enter>", efeito_hover)
    btn_salvar.bind("<Leave>", efeito_sair_hover)
    btn_salvar.pack(side="left", padx=5)

    btn_excluir = tk.Button(frame_botoes, text="Excluir Registro", bg="#dc2626", fg="white", command=excluir_registro)
    btn_excluir.original_bg = "#dc2626"
    configurar_estilo_botoes(btn_excluir)
    btn_excluir.bind("<Enter>", efeito_hover)
    btn_excluir.bind("<Leave>", efeito_sair_hover)
    btn_excluir.pack(side="left", padx=5)

    btn_gerar_excel = tk.Button(nova_janela, text="Preencher Dados Cliente/Processo", bg="#22c55e", fg="white", command=gerar_excel)
    btn_gerar_excel.original_bg = "#22c55e"
    configurar_estilo_botoes(btn_gerar_excel)
    btn_gerar_excel.bind("<Enter>", efeito_hover)
    btn_gerar_excel.bind("<Leave>", efeito_sair_hover)
    btn_gerar_excel.pack(pady=10)






def recurso_path(rel_path):
    try:
        base_path = sys._MEIPASS  # usado quando empacotado pelo PyInstaller
    except Exception:
        base_path = os.path.abspath(".")  # usado no modo desenvolvimento
    return os.path.join(base_path, rel_path)


def caminho_json():
    if getattr(sys, 'frozen', False):
        # Executável (.exe)
        base_path = os.path.dirname(sys.executable)
    else:
        # Modo desenvolvimento (.py)
        base_path = os.path.dirname(os.path.abspath(__file__))
    return os.path.join(base_path, "CAMINHO_JSON")

CAMINHO_JSON = caminho_json()


# Função para carregar os prazos salvos
def carregar_prazos():
    if not os.path.exists(CAMINHO_JSON):
        return []
    with open(CAMINHO_JSON, "r", encoding="utf-8") as f:
        return json.load(f)



# Função para salvar os prazos no JSON





# Função para carregar notificações
def carregar_notificacoes():
    tree_notificacoes.delete(*tree_notificacoes.get_children())
    try:
        prazos = carregar_prazos()
        for prazo in prazos:
            if prazo.get("notificado"):
                tree_notificacoes.insert("", "end", values=(
                    prazo.get("cliente", ""),
                    prazo.get("processo", ""),
                    prazo.get("tipo_prazo", ""),
                    prazo.get("data_para_notificar", ""),
                    prazo.get("registro_em", "")
                ))
    except Exception as e:
        messagebox.showerror("Erro", f"Erro ao carregar notificações:\n{e}")

# Função para salvar os prazos no JSON
def salvar_prazos(prazos):
    carregar_notificacoes()
    with open(CAMINHO_JSON, "w", encoding="utf-8") as f:
        json.dump(prazos, f, indent=4, ensure_ascii=False)


# Função para registrar novos prazos no JSON
# Ela será chamada dentro do gerar_excel()
def registrar_prazos_em_json():
    prazos_existentes = carregar_prazos()
    novos_registros = []
    for nome, processo, tipo, data, data_notificar, resp, pub, obs in registros:
        novo = {
            "cliente": nome,
            "processo": processo,
            "tipo_prazo": tipo,
            "data_fatal": data,
            "data_para_notificar": datetime.strptime(data_notificar.strip(), "%d/%m").strftime("%d/%m"),
            "lembrete": "FATAL",
            "lembrete": "FATAL",
            "notificado": False,
            "registro_em": datetime.now().strftime("%Y-%m-%d %H:%M")
        }
        novos_registros.append(novo)

    prazos_existentes.extend(novos_registros)
    salvar_prazos(prazos_existentes)

# Verifica se há prazos para hoje e avisa
# Pode ser chamado ao abrir o sistema ou com um botão próprio

def verificar_prazos_hoje():
    hoje = datetime.now().strftime("%d/%m")
    prazos = carregar_prazos()
    alertas = []

    for prazo in prazos:
        data_notificar = prazo.get("data_para_notificar", "").strip()
        if data_notificar == hoje:
            alertas.append(prazo)
            prazo["notificado"] = True

    if alertas:
        texto = "\n".join([
            "{} - {} - {} (FATAL)".format(p['cliente'], p['tipo_prazo'], p['processo'])
            for p in alertas
        ])
        messagebox.showinfo("🔔 Notificação", "Prazos configurados para hoje:\n\n{}".format(texto))

    salvar_prazos(prazos)
    carregar_notificacoes()




# Chamada da verificação automática ao abrir o app


print("Iniciando app...")

ARQUIVO_MODELO = "Planilha de prazos - atualizada.xlsx"

def recurso_path(rel_path):
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, rel_path)

IMAGEM_JURIDICA = recurso_path("justica.png")

PRAZOS = {
    "Direito Trabalhista - CLT": {
        "Agravo TST": 8,
        "Agravo de Instrumento em RR ou RO": 8,
        "Agravo de Petição": 8,
        "Contraminuta AI / Contrarrazões RR": 8,
        "Contrarrazões ao RR ou RO": 8,
        "Embargos TST": 8,
        "Embargos à Execução": 5,
        "Embargos de declaração": 5,
        "Impugnação à Sentença de Liquidação": 5,
        "Recurso de Revista": 8,
        "Recurso Ordinário Trabalhista": 8
    },
    "Direito Civil - CPC": {
        "Agravo de Instrumento": 15,
        "Apelação": 15,
        "Embargos de declaração": 5,
        "Recurso Especial": 15,
        "Recurso Extraordinário": 15,
        "Recurso Ordinário (Cível)": 15
    }
}

feriados_selecionados = []
registros = []

def calcular_prazo_util(data_str, dias):
    try:
        data_inicial = datetime.strptime(data_str, "%d/%m")
        data_inicial = data_inicial.replace(year=datetime.now().year)
        df = pd.date_range(data_inicial + timedelta(days=1), periods=90, freq='B')
        df_filtrado = [d for d in df if d.strftime("%d/%m") not in feriados_selecionados]
        if len(df_filtrado) < dias:
            return "Prazo ultrapassa os dias úteis disponíveis"
        termo_final = df_filtrado[dias - 1]
        return termo_final.strftime("%d/%m")
    except Exception as e:
        return f"Erro: {e}"

def exibir_calculo():
    data = publicacao_entry.get()
    ramo = ramo_var.get()
    tipo = tipo_var.get()
    if ramo in PRAZOS and tipo in PRAZOS[ramo]:
        prazo = PRAZOS[ramo][tipo]
        resultado = calcular_prazo_util(data, prazo)
        messagebox.showinfo("Resultado do Cálculo", f"{tipo}\nPrazo final: {resultado}")
    else:
        messagebox.showerror("Erro", "Selecione um tipo e ramo válido.")

def alternar_feriado():
    if checkbox_var.get():
        feriado_frame.pack(pady=5)
    else:
        feriado_frame.pack_forget()
        feriados_selecionados.clear()

def adicionar_feriado():
    inicio = feriado_inicio.get()
    fim = feriado_fim.get()
    try:
        data_inicio = datetime.strptime(inicio, "%d/%m")
        data_fim = datetime.strptime(fim, "%d/%m")
        ano_atual = datetime.now().year
        data_inicio = data_inicio.replace(year=ano_atual)
        data_fim = data_fim.replace(year=ano_atual)
        datas = pd.date_range(data_inicio, data_fim).to_pydatetime()
        for d in datas:
            feriados_selecionados.append(d.strftime("%d/%m"))
        feriado_inicio.delete(0, tk.END)
        feriado_fim.delete(0, tk.END)
        messagebox.showinfo("Feriado adicionado", f"Feriado(s) registrado(s): {inicio} até {fim}")
    except ValueError:
        messagebox.showerror("Erro", "Formato de data inválido. Use DD/MM.")


    criar_label("Aplicação desenvolvida por: Rodrigo Junqueira de Lima Siqueira").pack(side="bottom", pady=(10, 5))

def recurso_path(rel_path):
    try:
        base_path = sys._MEIPASS  # quando empacotado
    except Exception:
        base_path = os.path.abspath(".")  # durante desenvolvimento
    return os.path.join(base_path, rel_path)


IMAGEM_JURIDICA = os.path.join(os.path.dirname(__file__), "justica.png")

# Funcionalidade para o relátorio de prazos fatais em PDF

from fpdf import FPDF
from datetime import datetime
from tkinter import simpledialog, messagebox, filedialog
import os
import traceback

def gerar_relatorio_fatais_pdf():
    try:
        data_escolhida = simpledialog.askstring("Data", "Informe a data (DD/MM):")
        if not data_escolhida:
            return

        try:
            datetime.strptime(data_escolhida, "%d/%m")
        except ValueError:
            messagebox.showerror("Erro", "Data inválida. Use o formato DD/MM.")
            return

        prazos = carregar_prazos()
        prazos_filtrados = [p for p in prazos if p.get("data_fatal") == data_escolhida]

        if not prazos_filtrados:
            messagebox.showinfo("Sem prazos", "Nenhum prazo fatal encontrado para essa data.")
            return

        pdf = FPDF()
        pdf.add_page()
        pdf.add_font("DejaVu", "", "DejaVuSans.ttf", uni=True)
        pdf.add_font("DejaVu", "B", "DejaVuSans-Bold.ttf", uni=True)
        pdf.add_font("DejaVu", "I", "DejaVuSans-Oblique.ttf", uni=True)

        # Fundo escuro
        pdf.set_fill_color(15, 23, 42)
        pdf.rect(0, 0, 210, 297, 'F')

        # Imagem centralizada
        if os.path.exists("justica.png"):
            pdf.image("justica.png", x=(210 - 30) / 2, y=10, w=30)

        # Título
        pdf.set_y(45)
        pdf.set_text_color(255, 255, 255)
        pdf.set_font("DejaVu", "B", 14)
        pdf.cell(0, 10, "Sistema de Controle de Prazos Jurídicos", ln=True, align="C")

        # Texto principal
        pdf.set_font("DejaVu", "", 12)
        pdf.ln(10)
        pdf.cell(0, 10, f"Data do Relatório: {datetime.now().strftime('%d/%m/%Y')}", ln=True)
        pdf.cell(0, 10, f"Prazos Fatais – {data_escolhida}", ln=True)
        pdf.ln(5)

        for p in prazos_filtrados:
            cliente = p.get("cliente", "N/A")
            processo = p.get("processo", "N/A")
            tipo = p.get("tipo_prazo", "N/A")
            linha = f"{cliente} – {processo} – {tipo}"
            pdf.multi_cell(0, 10, linha)

        # Rodapé
        pdf.ln(10)
        pdf.set_font("DejaVu", "I", 10)
        pdf.set_text_color(200, 200, 200)
        pdf.cell(0, 10, "Relatório gerado automaticamente pelo sistema.", ln=True, align="C")
        pdf.cell(0, 10, "Desenvolvido por Rodrigo Junqueira de Lima Siqueira", ln=True, align="C")

        # Perguntar onde salvar
        nome_arquivo = filedialog.asksaveasfilename(
            defaultextension=".pdf",
            filetypes=[("PDF files", "*.pdf")],
            initialfile=f"Relatorio_Fatais_{data_escolhida.replace('/', '-')}.pdf",
            title="Salvar Relatório PDF"
        )

        if nome_arquivo:
            pdf.output(nome_arquivo)
            messagebox.showinfo("Sucesso", f"Relatório salvo como:\n{nome_arquivo}")
        else:
            messagebox.showinfo("Cancelado", "Salvamento do relatório foi cancelado.")

    except Exception as e:
        print("Erro ao gerar PDF:", traceback.format_exc())
        messagebox.showerror("Erro", f"Erro ao gerar PDF:\n{e}")








# Interface principal com abas no topo
janela = tk.Tk()
# === Estilo Global dos Botões ===
def configurar_estilo_botoes(widget):
    widget.configure(font=("Segoe UI", 9, "bold"), padx=4, pady=2, relief="raised", bd=2)

def efeito_hover(event):
    event.widget.config(bg="#3b82f6")  # Azul um pouco mais claro no hover

def efeito_sair_hover(event):
    event.widget.config(bg=event.widget.original_bg)  # Voltar para a cor original

janela.title("Sistema de Controle de Prazos")
janela.geometry("1150x720")
janela.configure(bg="#0f172a")

style = ttk.Style()
style.theme_use('default')
style.configure("TNotebook", background="#0f172a", borderwidth=0)
style.configure("TNotebook.Tab", background="#1e293b", foreground="white", font=("Segoe UI", 10, "bold"), padding=(20, 10))
style.map("TNotebook.Tab", background=[("selected", "#2563eb")], foreground=[("selected", "white")])

# Carregar ícones brancos
icon_inicio = tk.PhotoImage(file=recurso_path("icon_inicio.png")).subsample(40, 40)
icon_calc = tk.PhotoImage(file=recurso_path("icon_calc.png")).subsample(40, 40)
icon_notification = tk.PhotoImage(file=recurso_path("icon_notification.png")).subsample(40, 40)
icon_listagem = tk.PhotoImage(file=recurso_path("icon_listagem.png")).subsample(40, 40)
icon_config = tk.PhotoImage(file=recurso_path("icon_config.png")).subsample(40, 40)

abas = ttk.Notebook(janela, style="TNotebook")
abas.pack(fill="both", expand=True)

# === Aba Principal ===
aba_principal = tk.Frame(abas, bg="#0f172a")
abas.add(aba_principal, text="  Início  ", image=icon_inicio, compound="left")

# Imagem
if os.path.exists(IMAGEM_JURIDICA):
    try:
        imagem = Image.open(IMAGEM_JURIDICA)
        imagem = imagem.resize((140, 140), Image.LANCZOS)
        imagem_tk = ImageTk.PhotoImage(imagem)
        img_label = tk.Label(aba_principal, image=imagem_tk, bg="#0f172a")
        img_label.image = imagem_tk
        img_label.pack(pady=(80, 5), anchor="center")
    except Exception as e:
        print("Erro ao carregar imagem:", e)

# Título
tk.Label(aba_principal, text="Sistema de Controle de Prazos", font=("Segoe UI", 16, "bold"), bg="#0f172a", fg="white").pack(pady=(10, 5), anchor="center")

# Botão conferir prazos fatais
tk.Button(aba_principal, text="Conferir prazos fatais do dia", font=("Segoe UI", 11, "bold"), bg="#dc2626", fg="white", command=verificar_prazos_hoje).pack(pady=(10, 5), anchor="center")

# Botão Gerar Relátorio de Prazos Fatais (PDF)
tk.Button(aba_principal, text="Gerar Relatório de Prazos Fatais (PDF)", font=("Segoe UI", 11, "bold"), 
          bg="#dc2626", fg="white", command=gerar_relatorio_fatais_pdf).pack(pady=(5, 30), anchor="center")


# Botão preencher
tk.Button(aba_principal, text="Preencher Cliente / Processo", font=("Segoe UI", 12, "bold"), bg="#22c55e", fg="white", command=abrir_janela_preenchimento_multiplo).pack(pady=(10, 30), anchor="center")

# === Aba Calculadora ===
aba_calculadora = tk.Frame(abas, bg="#0f172a")
abas.add(aba_calculadora, text="  Calculadora de Prazos  ", image=icon_calc, compound="left")

# Campos da calculadora
tk.Label(aba_calculadora, text="Digite aqui a data de publicação", bg="#0f172a", fg="white", font=("Segoe UI", 16, "bold")).pack(pady=(100, 5), anchor="center")
publicacao_entry = tk.Entry(aba_calculadora, font=("Consolas", 14), width=20, justify="center", bd=2, relief="flat", bg="#1e293b", fg="#f8fafc", insertbackground="#f8fafc")
publicacao_entry.pack(pady=5, anchor="center")

ramo_var = tk.StringVar()
tipo_var = tk.StringVar()
ramo_menu = ttk.Combobox(aba_calculadora, textvariable=ramo_var, values=list(PRAZOS.keys()), state="readonly", width=40)
ramo_menu.pack(pady=2, anchor="center")

def atualizar_tipos(event):
    ramo = ramo_var.get()
    tipo_menu["values"] = sorted(PRAZOS.get(ramo, {}).keys())
    tipo_var.set("")

ramo_menu.bind("<<ComboboxSelected>>", atualizar_tipos)

tipo_menu = ttk.Combobox(aba_calculadora, textvariable=tipo_var, state="readonly", width=40)
tipo_menu.pack(pady=2, anchor="center")

# Botão calcular
tk.Button(aba_calculadora, text="Calcular Prazo Jurídico", font=("Segoe UI", 11, "bold"), bg="#2563eb", fg="white", command=exibir_calculo).pack(pady=10, anchor="center")

# Checkbox feriado
checkbox_var = tk.BooleanVar()
tk.Checkbutton(aba_calculadora, text="Ao longo do prazo existem feriados?", variable=checkbox_var, command=alternar_feriado, bg="#0f172a", fg="white", selectcolor="#0f172a", font=("Segoe UI", 11)).pack(pady=(5, 0), anchor="center")

feriado_frame = tk.Frame(aba_calculadora, bg="#0f172a")
feriado_inicio = tk.Entry(feriado_frame, font=("Consolas", 12), width=10, justify="center", bg="#1e293b", fg="#f8fafc")
feriado_inicio.grid(row=0, column=1, padx=5)
feriado_fim = tk.Entry(feriado_frame, font=("Consolas", 12), width=10, justify="center", bg="#1e293b", fg="#f8fafc")
feriado_fim.grid(row=0, column=3, padx=5)

adicionar_feriado_btn = tk.Button(feriado_frame, text="Adicionar Feriado", font=("Segoe UI", 10, "bold"), bg="#334155", fg="white", command=adicionar_feriado)
adicionar_feriado_btn.grid(row=0, column=4, padx=5)

# === Aba Notificações ===
aba_notificacoes = tk.Frame(abas, bg="#0f172a")
abas.add(aba_notificacoes, text="  Notificações do dia ", image=icon_notification, compound="left")
# === Aba Notificações Futuras ===
aba_notificacoes_futuras = tk.Frame(abas, bg="#0f172a")
abas.add(aba_notificacoes_futuras, text="  Notificações Futuras  ", image=icon_notification, compound="left")

tk.Label(aba_notificacoes_futuras, text="Notificações Agendadas", font=("Segoe UI", 14, "bold"),
         bg="#0f172a", fg="white").pack(pady=(20, 10))

colunas_futuras = ("Cliente", "Processo", "Tipo de Prazo", "Data Notificação Fatal", "Data", "Registro em")
tree_futuras = ttk.Treeview(aba_notificacoes_futuras, columns=colunas_futuras, show="headings", height=20)

for col in colunas_futuras:
    tree_futuras.heading(col, text=col)
    tree_futuras.column(col, width=130, anchor="center")

tree_futuras.pack(fill="both", padx=10, pady=(0, 10), expand=True)

def carregar_notificacoes_futuras():
    tree_futuras.delete(*tree_futuras.get_children())
    try:
        prazos = carregar_prazos()
        for prazo in prazos:
            if not prazo.get("notificado"):
                tree_futuras.insert("", "end", values=(
                    prazo.get("cliente", ""),
                    prazo.get("processo", ""),
                    prazo.get("tipo_prazo", ""),
                    prazo.get("data_para_notificar", ""),
                    prazo.get("data_fatal", ""),
                    prazo.get("registro_em", "")
                ))
    except Exception as e:
        messagebox.showerror("Erro", f"Erro ao carregar notificações futuras:\n{e}")

# Inicializa exibição
carregar_notificacoes_futuras()

tk.Label(aba_notificacoes, text="Histórico de Notificações", font=("Segoe UI", 14, "bold"),
         bg="#0f172a", fg="white").pack(pady=(20, 10))

# Treeview
colunas = ("Cliente", "Processo", "Tipo de Prazo", "Data Fatal", "Data de Registro")
tree_notificacoes = ttk.Treeview(aba_notificacoes, columns=colunas, show="headings", height=20)

for col in colunas:
    tree_notificacoes.heading(col, text=col)
    tree_notificacoes.column(col, width=150, anchor="center")

tree_notificacoes.pack(fill="both", padx=10, pady=(0, 10), expand=True)

def carregar_notificacoes():
    tree_notificacoes.delete(*tree_notificacoes.get_children())  # Limpa tabela
    try:
        prazos = carregar_prazos()
        for prazo in prazos:
            if prazo.get("notificado"):
                tree_notificacoes.insert("", "end", values=(
                    prazo.get("cliente", ""),
                    prazo.get("processo", ""),
                    prazo.get("tipo_prazo", ""),
                    prazo.get("data_para_notificar", ""),
                    prazo.get("registro_em", "")
                ))
    except Exception as e:
        messagebox.showerror("Erro", f"Erro ao carregar notificações:\n{e}")

def limpar_notificacoes():
    if messagebox.askyesno("Confirmação", "Deseja realmente limpar todas as notificações?"):
        try:
            prazos = carregar_prazos()
            prazos_filtrados = [p for p in prazos if not p.get("notificado")]
            salvar_prazos(prazos_filtrados)
            carregar_notificacoes()
            messagebox.showinfo("Sucesso", "Notificações removidas com sucesso.")
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao limpar notificações:\n{e}")

tk.Button(aba_notificacoes, text="Limpar Notificações", bg="#dc2626", fg="white",
          font=("Segoe UI", 10, "bold"), command=limpar_notificacoes).pack(pady=(0, 10))

# Inicializa exibição
carregar_notificacoes()




# === Aba Listagem ===
aba_listagem = tk.Frame(abas, bg="#0f172a")
abas.add(aba_listagem, text="  Listagem de Prazos  ", image=icon_listagem, compound="left")

tk.Label(aba_listagem, text="Listar prazos por data", font=("Segoe UI", 14, "bold"),
         bg="#0f172a", fg="white").pack(pady=(20, 5))

frame_data = tk.Frame(aba_listagem, bg="#0f172a")
frame_data.pack(pady=5)

tk.Label(frame_data, text="Data (DD/MM):", font=("Segoe UI", 10), bg="#0f172a", fg="white").pack(side="left", padx=5)
entry_data_busca = tk.Entry(frame_data, font=("Segoe UI", 11), width=10, justify="center")
entry_data_busca.pack(side="left", padx=5)
tk.Label(aba_listagem, text="⚠️ Dica: a busca considera a data de registro no sistema, não a data fatal.",
         font=("Segoe UI", 8), bg="#0f172a", fg="#facc15", justify="left").pack(pady=(2, 10), padx=10, anchor="w")

# Treeview
colunas_listagem = ("Cliente", "Processo", "Tipo de Prazo", "Data Fatal", "Registro em")
tree_listagem = ttk.Treeview(aba_listagem, columns=colunas_listagem, show="headings", height=20)

for col in colunas_listagem:
    tree_listagem.heading(col, text=col)
    tree_listagem.column(col, width=150, anchor="center")

tree_listagem.pack(fill="both", padx=10, pady=(10, 5), expand=True)

def buscar_prazos_por_data():
    data = entry_data_busca.get().strip()
    if not data:
        messagebox.showwarning("Aviso", "Insira uma data no formato DD/MM.")
        return
    try:
        datetime.strptime(data, "%d/%m")
    except ValueError:
        messagebox.showerror("Erro", "Data inválida. Use o formato DD/MM.")
        return

    tree_listagem.delete(*tree_listagem.get_children())
    prazos = carregar_prazos()
    encontrados = [p for p in prazos if p.get("data_fatal") == data]
    if not encontrados:
        messagebox.showinfo("Resultado", "Nenhum prazo encontrado para essa data.")
        return

    for p in encontrados:
        tree_listagem.insert("", "end", values=(
            p.get("cliente", ""),
            p.get("processo", ""),
            p.get("tipo_prazo", ""),
            p.get("data_para_notificar", ""),
            p.get("registro_em", "")
        ))

tk.Button(aba_listagem, text="Buscar", font=("Segoe UI", 10, "bold"),
          bg="#2563eb", fg="white", command=buscar_prazos_por_data).pack(pady=(0, 10))


# === Aba Configurações ===
aba_config = tk.Frame(abas, bg="#0f172a")
abas.add(aba_config, text="  Configurações  ", image=icon_config, compound="left")

tk.Label(aba_config, text="Configurações do Sistema", font=("Segoe UI", 14, "bold"),
         bg="#0f172a", fg="white").pack(pady=(20, 10))

def limpar_todos_os_prazos():
    resposta = messagebox.askyesnocancel(
        "⚠️ Exportar antes de apagar?",
        "Deseja exportar um backup dos prazos antes de apagar tudo?\n\nVocê poderá importá-los novamente pelo sistema."
    )

    if resposta is None:
        return  # Cancelado

    elif resposta:  # Sim, exportar antes de apagar
        try:
            prazos = carregar_prazos()
            if not prazos:
                messagebox.showinfo("Nada para exportar", "Não há prazos para exportar.")
                return

            nome_json = f"backup_prazos_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
            nome_zip = nome_json.replace(".json", ".zip")

            with open(nome_json, "w", encoding="utf-8") as f:
                json.dump(prazos, f, indent=4, ensure_ascii=False)

            with zipfile.ZipFile(nome_zip, "w", zipfile.ZIP_DEFLATED) as zipf:
                zipf.write(nome_json)

            os.remove(nome_json)  # limpa o json temporário

            messagebox.showinfo("Backup gerado", f"Backup salvo como:\n{nome_zip}\n\nVocê pode importá-lo pela aba Configurações.")
        except Exception as e:
            messagebox.showerror("Erro ao exportar", f"Falha ao gerar o backup:\n{e}")
            return  # cancela a limpeza caso falhe

    # Segunda confirmação
    confirmar = messagebox.askyesno(
        "⚠️ Confirmar exclusão",
        "Tem certeza que deseja apagar TODOS os prazos?\n\nEsta ação é irreversível."
    )

    if confirmar:
        try:
            salvar_prazos([])
            messagebox.showinfo("Sucesso", "Todos os prazos foram apagados com segurança.")
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao apagar: {e}")


def exportar_backup():
    local = filedialog.asksaveasfilename(defaultextension=".json", filetypes=[("JSON", "*.json")])
    if local:
        try:
            with open(local, "w", encoding="utf-8") as f:
                json.dump(carregar_prazos(), f, indent=4, ensure_ascii=False)
            messagebox.showinfo("Backup", f"Backup salvo com sucesso em:\n{local}")
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao exportar backup:\n{e}")

def importar_backup():
    caminho = filedialog.askopenfilename(filetypes=[("JSON", "*.json")])
    if caminho:
        try:
            with open(caminho, "r", encoding="utf-8") as f:
                dados = json.load(f)
            salvar_prazos(dados)
            messagebox.showinfo("Restauração", "Backup restaurado com sucesso.")
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao restaurar backup:\n{e}")

tk.Button(aba_config, text="Exportar Backup de Prazos", bg="#3b82f6", fg="white",
          font=("Segoe UI", 10, "bold"), command=exportar_backup).pack(pady=5)

tk.Button(aba_config, text="Importar Backup de Prazos", bg="#10b981", fg="white",
          font=("Segoe UI", 10, "bold"), command=importar_backup).pack(pady=5)

tk.Button(aba_config, text="Limpar Todos os Prazos", bg="#dc2626", fg="white",
          font=("Segoe UI", 10, "bold"), command=limpar_todos_os_prazos).pack(pady=5)

tk.Label(aba_config, text="Desenvolvido por Rodrigo Junqueira - Versão 1.0",
         bg="#0f172a", fg="#94a3b8", font=("Segoe UI", 9)).pack(side="bottom", pady=(10, 5))


# Funções placeholder para evitar erro de execução

# Rodapé
tk.Label(janela, text="Aplicação desenvolvida por: Rodrigo Junqueira de Lima Siqueira", font=("Segoe UI", 8), bg="#0f172a", fg="#64748b").pack(side="bottom", pady=(10, 5))

carregar_notificacoes()
verificar_prazos_hoje()
janela.mainloop()
