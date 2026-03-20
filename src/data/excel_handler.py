"""Gestão de arquivos Excel"""

from openpyxl import load_workbook
import os


class ExcelHandler:
    """Gerencia operações com arquivos Excel"""

    def __init__(self, caminho_modelo):
        self.caminho_modelo = caminho_modelo

    def arquivo_existe(self):
        """Verifica se arquivo modelo existe"""
        return os.path.exists(self.caminho_modelo)

    def preencher_prazos(self, registros, caminho_saida):
        """
        Preenche Excel com múltiplos registros de prazos
        
        Args:
            registros: Lista de tuplas (nome, processo, tipo, data, data_notificar, resp, pub, obs)
            caminho_saida: Caminho onde salvar o arquivo preenchido
            
        Returns:
            bool: True se sucesso, False caso contrário
        """
        try:
            wb = load_workbook(self.caminho_modelo)
            ws = wb.active
            bloco_index = 0

            for i, row in enumerate(ws.iter_rows(min_row=1)):
                if bloco_index >= len(registros):
                    break

                # Procura por linhas com "NOME" como marcador de bloco
                if row[0].value and str(row[0].value).strip().upper() == "NOME" and row[1].value in (None, ""):
                    nome, processo, tipo, data, data_notificar, resp, pub, obs = registros[bloco_index]

                    # Preenche dados
                    ws.cell(row=i + 1, column=2).value = nome
                    ws.cell(row=i + 2, column=2).value = processo
                    ws.cell(row=i + 3, column=2).value = tipo
                    ws.cell(row=i + 1, column=4).value = resp
                    ws.cell(row=i + 2, column=4).value = pub
                    ws.cell(row=i + 3, column=4).value = data
                    ws.cell(row=i + 1, column=6).value = obs

                    bloco_index += 1

            wb.save(caminho_saida)
            return True

        except Exception as e:
            print(f"Erro ao preencher Excel: {e}")
            return False
